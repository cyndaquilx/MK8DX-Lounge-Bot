import discord
from discord.ext import commands
from discord import app_commands

import openpyxl
import json

import API.post, API.get
import asyncio
import aiohttp

from constants import place_MMRs, channels, getRank, ranks, placementRoleID, player_role_ID

class Admin(commands.Cog):
    def __init__ (self, bot):
        self.bot = bot
        self.stopped = False

    #@commands.has_any_role("Administrator")
    #@commands.command(aliases=['pc'])
    async def placeEveryone(self, ctx):
        await ctx.send("working...")
        wb = openpyxl.load_workbook("s7Changes.xlsx", data_only=True)
        ws = wb["Sheet"]
        for i in range(10001, 17547):
            if(i % 100 == 0):
                await ctx.send(f"{i}/{17546}")
            name = ws[f"A{i}"].value
            mmr = ws[f"E{i}"].value
            success, player = await API.post.placePlayer(mmr, name)
            if success is False:
                print(f"{name} - {player}")
                continue
            await asyncio.sleep(0.05)
        await ctx.send("done")

    # old command when we completely remade roles, not used anymore
    #@commands.has_any_role("Administrator")
    #@commands.command(aliases=['all'])
    async def addAll(self, ctx):
        apiplayers = await API.get.getPlayerList()
        players = apiplayers['players']
        print('added:')
        for player in players:
            if 'discordId' not in player.keys():
                continue
            if 'mmr' not in player.keys():
                continue
            rank = getRank(player['mmr'])
            role = ctx.guild.get_role(ranks[rank]['roleid'])
            member = ctx.guild.get_member(int(player['discordId']))
            if member is None:
                continue
            for mrole in member.roles:
                if mrole.id == role.id:
                    continue
            await member.add_roles(role)
            print(f"{player['name']} - {role.name}")

    async def unlockdown(self, channel:discord.TextChannel):
        overwrite = channel.overwrites_for(channel.guild.default_role)
        overwrite.send_messages = None
        await channel.set_permissions(channel.guild.default_role, overwrite=overwrite)
        await channel.send("Unlocked " + channel.mention)

    # use this after all players have been placed on the website for new season
    @commands.has_any_role("Administrator")
    @commands.command()
    async def fixAllRoles(self, ctx):
        i = 0
        for member in ctx.guild.members:
            i+=1
            if i % 100 == 0:
                await ctx.send(f"{i}/{len(ctx.guild.members)}")
            playerRoles = []
            for role in member.roles:
                for rank in ranks.values():
                    if role.id == rank['roleid']:
                        playerRoles.append(role)
                    if role.id == placementRoleID:
                        playerRoles.append(role)

            player = await API.get.getPlayerFromDiscord(member.id)
            if player is None:
                for role in playerRoles:
                    try:
                        await member.remove_roles(role)
                    except Exception as e:
                        print(e)
                        continue
                continue
            
            if 'mmr' not in player.keys():
                role = member.guild.get_role(placementRoleID)
            else:
                rank = getRank(player['mmr'])
                role = member.guild.get_role(ranks[rank]['roleid'])

            for currRole in playerRoles:
                if currRole.id != role.id:
                    try:
                        await member.remove_roles(currRole)
                    except Exception as e:
                        print(e)
            if role not in playerRoles:
                try:
                    await member.add_roles(role)
                except Exception as e:
                    print(e)
        await ctx.send("done")

    #@commands.has_any_role("Administrator")
    #@commands.command()
    async def startseason(self, ctx):
        for channel in ctx.guild.channels:
            if channel.category_id in [445404698795573250, 876282435623608330, 1003118792794177568]:
                await self.unlockdown(channel)
        await ctx.send("All tier chats have been unlocked. ENJOY SEASON 7!! @everyone")
        
    #@commands.has_any_role("Administrator")
    #@commands.command()
    async def checkNumDiscords(self, ctx):
        apiplayers = await API.get.getPlayerList()
        players = apiplayers['players']
        count = 0
        for player in players:
            if 'discordId' in player.keys():
                count += 1
        await ctx.send(f"{count}/{len(players)} players have Discord accounts added, which is {(count/len(players)*100):.2f}% of the players in the database")
 
    #@commands.has_any_role("Administrator")
    #@commands.command()
    async def regmkc(self, ctx):
        await ctx.send("working...")
        wb = openpyxl.load_workbook("tourney bonus.xlsx", data_only=True)
        ws = wb["Sheet1"]
        base_url = "https://mariokartcentral.com/mkc/api/registry/players/"
        connector=aiohttp.TCPConnector(ssl=False)
        async with aiohttp.ClientSession(connector=connector) as session:
            for i in range(2, 745):
                full_url = base_url + str(ws[f"A{i}"].value)
                async with session.get(full_url) as resp:
                    info = await resp.json()
                    ws[f"E{i}"] = info["user_id"]
                    if i % 100 == 0:
                        await ctx.send(i)
                    await asyncio.sleep(0.1)
        wb.save("tourney bonus.xlsx")
        await ctx.send("done")

    #@commands.has_any_role("Administrator")
    #@commands.command()
    async def getname(self, ctx):
        await ctx.send("working...")
        wb = openpyxl.load_workbook("tourney bonus.xlsx", data_only=True)
        ws = wb["Sheet1"]
        for i in range(2, 745):
            mkcid = ws[f"E{i}"].value
            p = await API.get.getPlayerFromMKC(mkcid)
            if p is not None:
                ws[f"F{i}"] = p['name']
            if i % 100 == 0:
                await ctx.send(i)
            await asyncio.sleep(0.1)
        wb.save("tourney bonus.xlsx")
        await ctx.send("done")

    #@commands.has_any_role("Administrator")
    #@commands.command()
    async def givebonuses(self, ctx):
        wb = openpyxl.load_workbook("tourney bonus.xlsx", data_only=True)
        ws = wb["Sheet1"]
        for i in range(429, 737):
            loungename = ws[f"F{i}"].value
            bonus = ws[f"D{i}"].value
            p = await API.get.getPlayer(loungename)
            a, b = await API.post.createBonus(loungename, bonus)
            if a is False:
                await ctx.send(f"{loungename}: {b}")
            else:
                ws[f"G{i}"] = "yes"
                mmr1 = p["mmr"]
                mmr2 = mmr1 + bonus
                ws[f"H{i}"] = mmr1
                ws[f"I{i}"] = mmr2
                oldRank = getRank(mmr1)
                newRank = getRank(mmr2)
                if oldRank != newRank:
                    ws[f"J{i}"] = "yes"
                    if 'discordId' in p.keys():
                        member = ctx.guild.get_member(int(p['discordId']))
                        oldRole = ctx.guild.get_role(ranks[oldRank]["roleid"])
                        newRole = ctx.guild.get_role(ranks[newRank]["roleid"])
                        if member is not None and oldRole is not None and newRole is not None:
                            if oldRole in member.roles:
                                await member.remove_roles(oldRole)
                            if newRole not in member.roles:
                                await member.add_roles(newRole)
                            ws[f"K{i}"] = "yes"
            if i % 100 == 0:
                await ctx.send(i)
            await asyncio.sleep(0.1)
        wb.save("tourney bonus.xlsx")
        await ctx.send("done")

    @commands.command()
    async def countchannels(self, ctx):
        count = 0
        for channel in ctx.guild.channels:
            count += 1
        await ctx.send(count)

    @commands.has_any_role("Administrator")
    @commands.command()
    async def give_player_role(self, ctx):
        self.stopped = False
        player_role = ctx.guild.get_role(player_role_ID)
        if player_role is None:
            await ctx.send("Player role could not be found")
            return
        rank_role_ids = [ranks[rank]['roleid'] for rank in ranks.keys()]
        rank_role_ids.append(placementRoleID)
        num_roles_given = 0
        await ctx.send("Working...")
        for i, member in enumerate(ctx.guild.members):
            if self.stopped is True:
                self.stopped = False
                await ctx.send(f"Stopped giving player role; checked {i}/{len(ctx.guild.members)} members and gave {num_roles_given} player roles")
                return
            role_ids = [r.id for r in member.roles]
            if player_role_ID in role_ids:
                continue
            for role in rank_role_ids:
                if role in role_ids:
                    await member.add_roles(player_role)
                    num_roles_given += 1
                    if num_roles_given % 100 == 0:
                        await ctx.send(f"Given {num_roles_given} player roles so far")
                    continue
        await ctx.send("Everyone who is supposed to have the Player role should now have it.")

    @commands.has_any_role("Administrator")
    @commands.command()
    async def stop_player_role(self, ctx):
        self.stopped = True

    @commands.command()
    @commands.is_owner()
    async def sync_server(self, ctx):
        await self.bot.tree.sync(guild=discord.Object(id=ctx.guild.id))
        await ctx.send("synced")

    @commands.command()
    @commands.is_owner()
    async def sync(self, ctx):
        await self.bot.tree.sync()
        await ctx.send("synced")
    
    @app_commands.command()
    @app_commands.guilds(445404006177570829)
    @app_commands.checks.has_permissions(administrator=True)
    async def add_admin_role(self, interaction:discord.Interaction, role:discord.Role):
        if str(interaction.guild_id) not in self.bot.server_config["admin_roles"].keys():
            self.bot.server_config["admin_roles"][str(interaction.guild_id)] = []
        roles_list = self.bot.server_config["admin_roles"][str(interaction.guild_id)]
        if role.id in roles_list:
            await interaction.response.send_message("This role already has admin permissions in this server")
        else:
            roles_list.append(role.id)
            with open('./server_config.json', 'w', encoding='utf-8') as f:
                json.dump(self.bot.server_config, f, ensure_ascii=False, indent=4)
            await interaction.response.send_message("Successfully added admin perms for this role")

    @app_commands.command()
    @app_commands.guilds(445404006177570829)
    @app_commands.checks.has_permissions(administrator=True)
    async def add_staff_role(self, interaction:discord.Interaction, role:discord.Role):
        if str(interaction.guild_id) not in self.bot.server_config["staff_roles"].keys():
            self.bot.server_config["staff_roles"][str(interaction.guild_id)] = []
        roles_list = self.bot.server_config["staff_roles"][str(interaction.guild_id)]
        if role.id in roles_list:
            await interaction.response.send_message("This role already has staff permissions in this server")
        else:
            roles_list.append(role.id)
            with open('./server_config.json', 'w', encoding='utf-8') as f:
                json.dump(self.bot.server_config, f, ensure_ascii=False, indent=4)
            await interaction.response.send_message("Successfully added staff perms for this role")

    @app_commands.command()
    @app_commands.guilds(445404006177570829)
    @app_commands.checks.has_permissions(administrator=True)
    async def add_reporter_role(self, interaction:discord.Interaction, role:discord.Role):
        if str(interaction.guild_id) not in self.bot.server_config["reporter_roles"].keys():
            self.bot.server_config["reporter_roles"][str(interaction.guild_id)] = []
        roles_list = self.bot.server_config["reporter_roles"][str(interaction.guild_id)]
        if role.id in roles_list:
            await interaction.response.send_message("This role already has reporter permissions in this server")
        else:
            roles_list.append(role.id)
            with open('./server_config.json', 'w', encoding='utf-8') as f:
                json.dump(self.bot.server_config, f, ensure_ascii=False, indent=4)
            await interaction.response.send_message("Successfully added reporter perms for this role")

    @app_commands.command()
    @app_commands.guilds(445404006177570829)
    @app_commands.checks.has_permissions(administrator=True)
    async def add_name_restricted_role(self, interaction:discord.Interaction, role:discord.Role):
        if str(interaction.guild_id) not in self.bot.server_config["name_restricted_roles"].keys():
            self.bot.server_config["name_restricted_roles"][str(interaction.guild_id)] = []
        roles_list = self.bot.server_config["name_restricted_roles"][str(interaction.guild_id)]
        if role.id in roles_list:
            await interaction.response.send_message("This role already has name restricted permissions in this server")
        else:
            roles_list.append(role.id)
            with open('./server_config.json', 'w', encoding='utf-8') as f:
                json.dump(self.bot.server_config, f, ensure_ascii=False, indent=4)
            await interaction.response.send_message("Successfully added name restricted perms for this role")

    @app_commands.command()
    @app_commands.guilds(445404006177570829)
    @app_commands.checks.has_permissions(administrator=True)
    async def add_chat_restricted_role(self, interaction:discord.Interaction, role:discord.Role):
        if str(interaction.guild_id) not in self.bot.server_config["chat_restricted_roles"].keys():
            self.bot.server_config["chat_restricted_roles"][str(interaction.guild_id)] = []
        roles_list = self.bot.server_config["chat_restricted_roles"][str(interaction.guild_id)]
        if role.id in roles_list:
            await interaction.response.send_message("This role already has chat restricted permissions in this server")
        else:
            roles_list.append(role.id)
            with open('./server_config.json', 'w', encoding='utf-8') as f:
                json.dump(self.bot.server_config, f, ensure_ascii=False, indent=4)
            await interaction.response.send_message("Successfully added chat restricted perms for this role")


    
async def setup(bot):
    await bot.add_cog(Admin(bot))
